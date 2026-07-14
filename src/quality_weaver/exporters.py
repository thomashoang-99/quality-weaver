import os
import re
import tempfile
from dataclasses import dataclass
from pathlib import Path
from zipfile import BadZipFile

import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import ValidationError

from quality_weaver.io import atomic_write_text
from quality_weaver.models import ApprovalStatus, TestCase, TestCaseDocument
from quality_weaver.profiles import Profile, WorkbookProfile
from quality_weaver.testcases import render_testcases_markdown
from quality_weaver.workspace import StateError, Workspace


@dataclass(frozen=True)
class ExportFinding:
    code: str
    message: str
    artifact_id: str
    blocking: bool = True


@dataclass(frozen=True)
class ExportResult:
    path: Path
    case_count: int


class ExportError(RuntimeError):
    """One or more deterministic export checks failed."""

    def __init__(self, *findings: ExportFinding) -> None:
        self.findings = tuple(sorted(findings, key=lambda item: (item.code, item.artifact_id)))
        super().__init__("; ".join(f"{item.code}: {item.message}" for item in self.findings))


def export_markdown(
    workspace: Workspace,
    document: TestCaseDocument,
    profile: Profile,
    output_path: Path,
    *,
    protected_inputs: tuple[Path, ...],
) -> ExportResult:
    """Atomically export canonical Markdown after all approval checks pass."""
    _ensure_ready(workspace, document)
    _ensure_format(profile, "markdown")
    _ensure_distinct_output(output_path, (*protected_inputs, *_profile_resources(profile)))
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_text(output_path, render_testcases_markdown(document))
    except OSError as error:
        raise _error("EXPORT_WRITE_FAILED", str(error), str(output_path)) from error
    return ExportResult(path=output_path, case_count=len(document.cases))


def export_excel(
    workspace: Workspace,
    document: TestCaseDocument,
    profile: Profile,
    *,
    workbook_kind: str,
    output_directory: Path,
    project: str,
    artifact: str,
    protected_inputs: tuple[Path, ...],
) -> ExportResult:
    """Render a verified workbook through a self-contained profile template."""
    _ensure_ready(workspace, document)
    _ensure_format(profile, "excel")
    workbook_profile = profile.workbooks.get(workbook_kind)
    if workbook_profile is None:
        raise _error(
            "EXPORT_WORKBOOK_UNKNOWN",
            f"unknown workbook kind: {workbook_kind}",
            workbook_kind,
        )
    _ensure_filename_value(project, "project")
    _ensure_filename_value(artifact, "artifact")
    try:
        validated_workbook = WorkbookProfile.model_validate(workbook_profile.model_dump())
        filename = validated_workbook.filename.format(project=project, artifact=artifact)
    except (AttributeError, IndexError, KeyError, ValidationError, ValueError) as error:
        raise _error(
            "EXPORT_FILENAME_INVALID",
            "filename policy must use plain project and artifact fields",
            workbook_profile.filename,
        ) from error
    if Path(filename).name != filename or Path(filename).suffix.lower() != ".xlsx":
        raise _error(
            "EXPORT_FILENAME_INVALID",
            "filename policy must produce one .xlsx name",
            filename,
        )

    output_path = output_directory / filename
    template_path = workbook_profile.template_path(profile.root)
    _ensure_distinct_output(output_path, (*protected_inputs, *_profile_resources(profile)))
    workbook = _load_template(template_path, workbook_profile)
    _clear_mapped_data(workbook, workbook_profile)
    _write_cases(workbook, workbook_profile, document)
    _fill_organization(workbook, profile, project)
    _verify_case_count(workbook, workbook_profile, document)
    _save_workbook_atomic(workbook, output_path, workbook_profile, document)
    return ExportResult(path=output_path, case_count=len(document.cases))


def _ensure_ready(workspace: Workspace, document: TestCaseDocument) -> None:
    try:
        workspace.ensure_export_ready()
    except StateError as error:
        raise _error("EXPORT_GATES_NOT_APPROVED", str(error), "workspace") from error
    if document.status is not ApprovalStatus.APPROVED:
        raise _error(
            "EXPORT_CASES_NOT_APPROVED",
            "test-case document must be approved before export",
            "testcases",
        )


def _ensure_format(profile: Profile, output_format: str) -> None:
    if output_format not in profile.formats:
        raise _error(
            "EXPORT_FORMAT_UNSUPPORTED",
            f"profile {profile.name} does not support {output_format}",
            profile.name,
        )


def _ensure_distinct_output(output_path: Path, protected_inputs: tuple[Path, ...]) -> None:
    resolved_output = output_path.resolve()
    if resolved_output in {path.resolve() for path in protected_inputs}:
        raise _error(
            "EXPORT_OUTPUT_COLLISION",
            "resolved output path collides with a protected input",
            str(output_path),
        )


def _profile_resources(profile: Profile) -> tuple[Path, ...]:
    templates = tuple(
        workbook.template_path(profile.root) for workbook in profile.workbooks.values()
    )
    return (profile.root / "profile.yaml", *templates)


def _ensure_filename_value(value: str, field: str) -> None:
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._ -]*", value) or value in {".", ".."}:
        raise _error(
            "EXPORT_FILENAME_VALUE_INVALID",
            f"{field} contains characters unsafe for a filename",
            field,
        )


def _load_template(path: Path, profile: WorkbookProfile) -> Workbook:
    try:
        workbook = openpyxl.load_workbook(path, data_only=False)
    except (OSError, InvalidFileException, ValueError) as error:
        raise _error("EXPORT_TEMPLATE_INVALID", str(error), str(path)) from error
    missing = sorted(set(profile.required_sheets) - set(workbook.sheetnames))
    if missing:
        raise _error(
            "EXPORT_TEMPLATE_SHEET_MISSING",
            f"template is missing required sheets: {', '.join(missing)}",
            str(path),
        )
    sheet = workbook[profile.sheet]
    mismatches = [
        field
        for field, expected in profile.headers.model_dump().items()
        if sheet.cell(
            row=profile.header_row,
            column=getattr(profile.columns, field),
        ).value
        != expected
    ]
    if mismatches:
        raise _error(
            "EXPORT_TEMPLATE_MAPPING_INVALID",
            f"template headers do not match declared mappings: {', '.join(sorted(mismatches))}",
            str(path),
        )
    return workbook


def _write_cases(
    workbook: Workbook, profile: WorkbookProfile, document: TestCaseDocument
) -> None:
    sheet = workbook[profile.sheet]
    for offset, test_case in enumerate(sorted(document.cases, key=lambda case: case.id)):
        row = profile.first_row + offset
        values = _case_values(test_case)
        for field, column in profile.columns.model_dump().items():
            cell = sheet.cell(row=row, column=column, value=values[field])
            cell.data_type = "s"


def _clear_mapped_data(workbook: Workbook, profile: WorkbookProfile) -> None:
    sheet = workbook[profile.sheet]
    mapped_columns = set(profile.columns.model_dump().values())
    vertical_merges = [
        str(cell_range)
        for cell_range in sheet.merged_cells.ranges
        if cell_range.max_row >= profile.first_row
        and cell_range.min_row != cell_range.max_row
        and any(
            cell_range.min_col <= column <= cell_range.max_col for column in mapped_columns
        )
    ]
    for cell_range in vertical_merges:
        sheet.unmerge_cells(cell_range)
    for row in range(profile.first_row, sheet.max_row + 1):
        for column in mapped_columns:
            cell = sheet.cell(row=row, column=column)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _case_values(test_case: TestCase) -> dict[str, str]:
    return {
        "id": test_case.id,
        "title": test_case.title,
        "traceability": (
            f"Outline: {test_case.outline_id}\n"
            f"Coverage: {', '.join(sorted(test_case.coverage_ids))}\n"
            f"Priority: {test_case.priority}\n"
            f"Tags: {', '.join(sorted(test_case.tags)) if test_case.tags else 'None.'}"
        ),
        "preconditions": (
            f"Preconditions:\n{_numbered(test_case.preconditions)}\n"
            f"Test Data:\n{_numbered(test_case.test_data)}"
        ),
        "steps": _numbered([step.action for step in test_case.steps]),
        "expected": _numbered([step.expected for step in test_case.steps]),
    }


def _numbered(values: list[str]) -> str:
    if not values:
        return "None."
    return "\n".join(f"{number}. {value}" for number, value in enumerate(values, start=1))


def _fill_organization(workbook: Workbook, profile: Profile, project: str) -> None:
    metadata = profile.organization
    if metadata is not None:
        workbook["Overview"][metadata.project_cell] = project


def _verify_case_count(
    workbook: Workbook, profile: WorkbookProfile, document: TestCaseDocument
) -> None:
    sheet = workbook[profile.sheet]
    ids = [
        sheet.cell(row=row, column=profile.columns.id).value
        for row in range(profile.first_row, sheet.max_row + 1)
        if sheet.cell(row=row, column=profile.columns.id).value is not None
    ]
    expected = [case.id for case in sorted(document.cases, key=lambda case: case.id)]
    if ids != expected:
        raise _error(
            "EXPORT_CASE_COUNT_MISMATCH",
            f"workbook case IDs do not match document: expected {len(expected)}",
            profile.sheet,
        )


def _save_workbook_atomic(
    workbook: Workbook,
    output_path: Path,
    profile: WorkbookProfile,
    document: TestCaseDocument,
) -> None:
    temporary_path: Path | None = None
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        descriptor, temporary_name = tempfile.mkstemp(
            dir=output_path.parent,
            prefix=f".{output_path.name}.",
            suffix=".tmp.xlsx",
        )
        temporary_path = Path(temporary_name)
        os.close(descriptor)
        workbook.save(temporary_path)
        verified_workbook = openpyxl.load_workbook(temporary_path, data_only=False)
        try:
            _verify_case_count(verified_workbook, profile, document)
        finally:
            verified_workbook.close()
        temporary_path.replace(output_path)
    except ExportError:
        raise
    except (BadZipFile, OSError, InvalidFileException, KeyError, ValueError) as error:
        raise _error("EXPORT_WRITE_FAILED", str(error), str(output_path)) from error
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def _error(code: str, message: str, artifact_id: str) -> ExportError:
    return ExportError(ExportFinding(code=code, message=message, artifact_id=artifact_id))
