import shutil
from hashlib import sha256
from pathlib import Path

import openpyxl
import pytest
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from typer.testing import CliRunner

from quality_weaver import exporters as exporter_module
from quality_weaver.cli import app
from quality_weaver.exporters import ExportError, export_excel, export_markdown
from quality_weaver.models import (
    ApprovalStatus,
)
from quality_weaver.models import (
    TestCase as CaseModel,
)
from quality_weaver.models import (
    TestCaseDocument as CaseDocument,
)
from quality_weaver.models import (
    TestStep as StepModel,
)
from quality_weaver.profiles import Profile
from quality_weaver.testcases import render_testcases_markdown
from quality_weaver.workspace import Stage, Workspace

PROFILES_ROOT = Path("profiles")


def copied_profile(name: str, destination: Path) -> Profile:
    profiles_root = destination / "profiles"
    shutil.copytree(PROFILES_ROOT / name, profiles_root / name)
    return Profile.load(name, profiles_root)


def document(*, status: ApprovalStatus = ApprovalStatus.APPROVED) -> CaseDocument:
    return CaseDocument(
        status=status,
        cases=[
            CaseModel(
                id="TC-002",
                title="Session timeout",
                outline_id="OUT-002",
                coverage_ids=["COV-002"],
                preconditions=["User is signed in"],
                test_data=[],
                steps=[StepModel(action="Wait for timeout", expected="Session expires")],
                priority="medium",
                tags=["session"],
            ),
            CaseModel(
                id="TC-001",
                title="Reject empty email",
                outline_id="OUT-001",
                coverage_ids=["COV-001", "COV-003"],
                preconditions=["User is on login", "Form is empty"],
                test_data=["email = empty"],
                steps=[
                    StepModel(action="Submit form", expected="Validation appears"),
                    StepModel(action="Inspect email", expected="Email stays empty"),
                ],
                priority="high",
                tags=["login"],
            ),
        ],
    )


def ready_workspace(tmp_path: Path) -> Workspace:
    workspace = Workspace.init(tmp_path)
    for stage in Stage:
        workspace.approve(stage)
    return workspace


def sha256_file(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()


@pytest.mark.parametrize("approved_count", [0, 1, 2])
def test_export_is_blocked_until_all_three_workspace_gates_are_approved(
    tmp_path: Path, approved_count: int
) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = Workspace.init(project)
    for stage in list(Stage)[:approved_count]:
        workspace.approve(stage)

    with pytest.raises(ExportError) as raised:
        export_markdown(
            workspace,
            document(),
            Profile.load("generic", PROFILES_ROOT),
            tmp_path / "out" / "testcases.md",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_GATES_NOT_APPROVED"
    assert not (tmp_path / "out" / "testcases.md").exists()


def test_export_requires_approved_document_and_compatible_format(tmp_path: Path) -> None:
    workspace = ready_workspace(tmp_path)
    generic = Profile.load("generic", PROFILES_ROOT)

    with pytest.raises(ExportError) as unapproved:
        export_markdown(
            workspace,
            document(status=ApprovalStatus.DRAFT),
            generic,
            tmp_path / "draft.md",
            protected_inputs=(),
        )
    with pytest.raises(ExportError) as incompatible:
        export_excel(
            workspace,
            document(),
            generic,
            workbook_kind="ut",
            output_directory=tmp_path,
            project="Demo",
            artifact="Login",
            protected_inputs=(),
        )

    assert unapproved.value.findings[0].code == "EXPORT_CASES_NOT_APPROVED"
    assert incompatible.value.findings[0].code == "EXPORT_FORMAT_UNSUPPORTED"


def test_generic_markdown_export_is_canonical_byte_for_byte_and_atomic(tmp_path: Path) -> None:
    workspace = ready_workspace(tmp_path)
    output = tmp_path / "exports" / "testcases.md"

    result = export_markdown(
        workspace,
        document(),
        Profile.load("generic", PROFILES_ROOT),
        output,
        protected_inputs=(),
    )

    assert result.path == output
    assert result.case_count == 2
    assert output.read_bytes() == render_testcases_markdown(document()).encode("utf-8")
    assert list(output.parent.glob(".testcases.md.*.tmp")) == []


def test_resolved_output_collision_is_rejected_before_write(tmp_path: Path) -> None:
    workspace = ready_workspace(tmp_path)
    cases_path = tmp_path / "cases.yaml"
    cases_path.write_text(document().model_dump_json(), encoding="utf-8")

    with pytest.raises(ExportError) as raised:
        export_markdown(
            workspace,
            document(),
            Profile.load("generic", PROFILES_ROOT),
            tmp_path / "alias" / ".." / "cases.yaml",
            protected_inputs=(cases_path,),
        )

    assert raised.value.findings[0].code == "EXPORT_OUTPUT_COLLISION"
    assert cases_path.read_text(encoding="utf-8") == document().model_dump_json()


def test_markdown_export_always_protects_every_profile_resource(tmp_path: Path) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = copied_profile("company-legacy", tmp_path)
    template = profile.workbooks["ut"].template_path(profile.root)
    before = template.read_bytes()

    with pytest.raises(ExportError) as raised:
        export_markdown(
            workspace,
            document(),
            profile,
            template,
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_OUTPUT_COLLISION"
    assert template.read_bytes() == before


def test_excel_export_protects_templates_for_unselected_workbook_kinds(
    tmp_path: Path,
) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = copied_profile("company-legacy", tmp_path)
    it_template = profile.workbooks["it"].template_path(profile.root)
    before = it_template.read_bytes()
    ut = profile.workbooks["ut"].model_copy(
        update={"filename": "{project}{artifact}.xlsx"}
    )
    profile = profile.model_copy(update={"workbooks": {**profile.workbooks, "ut": ut}})

    with pytest.raises(ExportError) as raised:
        export_excel(
            workspace,
            document(),
            profile,
            workbook_kind="ut",
            output_directory=it_template.parent,
            project="IT_",
            artifact="TestCase",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_OUTPUT_COLLISION"
    assert it_template.read_bytes() == before


@pytest.mark.parametrize(
    "kind, expected_name",
    [
        ("ut", "Demo_Login_Test Case UT.xlsx"),
        ("it", "Demo_Login_Test Case IT.xlsx"),
    ],
)
def test_excel_export_preserves_template_maps_cases_and_verifies_count(
    tmp_path: Path, kind: str, expected_name: str
) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT)
    template = profile.workbooks[kind].template_path(profile.root)
    before = sha256_file(template)

    result = export_excel(
        workspace,
        document(),
        profile,
        workbook_kind=kind,
        output_directory=tmp_path / "exports",
        project="Demo",
        artifact="Login",
        protected_inputs=(),
    )

    assert result.path.name == expected_name
    assert result.case_count == 2
    assert sha256_file(template) == before
    workbook = openpyxl.load_workbook(result.path, data_only=False)
    sheet = workbook["Testcase"]
    assert sheet["A16"].value == "TC-001"
    assert sheet["B16"].value == "Reject empty email"
    assert sheet["D16"].value == (
        "Outline: OUT-001\nCoverage: COV-001, COV-003\nPriority: high\nTags: login"
    )
    assert sheet["G16"].value == (
        "Preconditions:\n1. User is on login\n2. Form is empty\n"
        "Test Data:\n1. email = empty"
    )
    assert sheet["I16"].value == "1. Submit form\n2. Inspect email"
    assert sheet["J16"].value == "1. Validation appears\n2. Email stays empty"
    assert sheet["A17"].value == "TC-002"
    assert sheet["G17"].value == (
        "Preconditions:\n1. User is signed in\nTest Data:\nNone."
    )
    assert list(result.path.parent.glob(f".{expected_name}.*.tmp")) == []


def test_excel_clears_stale_mapped_rows_and_verifies_all_ids_after_save(
    tmp_path: Path, monkeypatch
) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT)
    original_load = exporter_module.openpyxl.load_workbook
    load_count = 0

    def load_with_stale_case(*args, **kwargs):
        nonlocal load_count
        workbook = original_load(*args, **kwargs)
        load_count += 1
        if load_count == 1:
            workbook["Testcase"]["A30"] = "TC-OLD"
            workbook["Testcase"]["B30"] = "Stale title"
        return workbook

    monkeypatch.setattr(exporter_module.openpyxl, "load_workbook", load_with_stale_case)

    result = export_excel(
        workspace,
        document(),
        profile,
        workbook_kind="ut",
        output_directory=tmp_path,
        project="Demo",
        artifact="Clean",
        protected_inputs=(),
    )

    sheet = original_load(result.path, data_only=False)["Testcase"]
    ids = [sheet.cell(row=row, column=1).value for row in range(16, sheet.max_row + 1)]
    assert [value for value in ids if value is not None] == ["TC-001", "TC-002"]
    assert load_count >= 2


def test_excel_reloads_temporary_workbook_before_atomic_publish(
    tmp_path: Path, monkeypatch
) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT)
    original_save = OpenpyxlWorkbook.save

    def save_then_corrupt(workbook, filename) -> None:
        original_save(workbook, filename)
        with Path(filename).open("rb") as temporary_file:
            corrupted = openpyxl.load_workbook(temporary_file, data_only=False)
        corrupted["Testcase"]["A16"] = "TC-CORRUPTED"
        original_save(corrupted, filename)

    monkeypatch.setattr(OpenpyxlWorkbook, "save", save_then_corrupt)
    output = tmp_path / "Demo_Corrupt_Test Case UT.xlsx"

    with pytest.raises(ExportError) as raised:
        export_excel(
            workspace,
            document(),
            profile,
            workbook_kind="ut",
            output_directory=tmp_path,
            project="Demo",
            artifact="Corrupt",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_CASE_COUNT_MISMATCH"
    assert not output.exists()


def test_excel_rejects_unknown_workbook_missing_sheet_and_unsafe_filename(
    tmp_path: Path,
) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT)

    with pytest.raises(ExportError) as unknown:
        export_excel(
            workspace,
            document(),
            profile,
            workbook_kind="e2e",
            output_directory=tmp_path,
            project="Demo",
            artifact="Login",
            protected_inputs=(),
        )
    with pytest.raises(ExportError) as unsafe:
        export_excel(
            workspace,
            document(),
            profile,
            workbook_kind="ut",
            output_directory=tmp_path,
            project="../Demo",
            artifact="Login",
            protected_inputs=(),
        )

    assert unknown.value.findings[0].code == "EXPORT_WORKBOOK_UNKNOWN"
    assert unsafe.value.findings[0].code == "EXPORT_FILENAME_VALUE_INVALID"

    workbook_profile = profile.workbooks["ut"].model_copy(
        update={"required_sheets": ("Overview", "Testcase", "Missing")}
    )
    broken = profile.model_copy(update={"workbooks": {"ut": workbook_profile}})
    with pytest.raises(ExportError) as missing:
        export_excel(
            workspace,
            document(),
            broken,
            workbook_kind="ut",
            output_directory=tmp_path,
            project="Demo",
            artifact="Login",
            protected_inputs=(),
        )
    assert missing.value.findings[0].code == "EXPORT_TEMPLATE_SHEET_MISSING"


def test_excel_rejects_template_whose_declared_column_header_changed(
    tmp_path: Path, monkeypatch
) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT)
    original_load = exporter_module.openpyxl.load_workbook

    def load_with_wrong_header(*args, **kwargs):
        workbook = original_load(*args, **kwargs)
        workbook["Testcase"]["A12"] = "Unexpected ID header"
        return workbook

    monkeypatch.setattr(exporter_module.openpyxl, "load_workbook", load_with_wrong_header)

    with pytest.raises(ExportError) as raised:
        export_excel(
            workspace,
            document(),
            profile,
            workbook_kind="ut",
            output_directory=tmp_path,
            project="Demo",
            artifact="Login",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_TEMPLATE_MAPPING_INVALID"


def test_excel_writes_user_text_as_literal_cells_not_formulas(tmp_path: Path) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT)
    cases = document()
    cases.cases[1].title = "=1+1"

    result = export_excel(
        workspace,
        cases,
        profile,
        workbook_kind="ut",
        output_directory=tmp_path,
        project="Demo",
        artifact="Formula",
        protected_inputs=(),
    )

    cell = openpyxl.load_workbook(result.path, data_only=False)["Testcase"]["B16"]
    assert cell.value == "=1+1"
    assert cell.data_type == "s"


def test_invalid_filename_policy_at_export_boundary_is_typed(tmp_path: Path) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT)
    ut = profile.workbooks["ut"].model_copy(
        update={"filename": "{project}_{missing}.xlsx"}
    )
    profile = profile.model_copy(update={"workbooks": {"ut": ut}})

    with pytest.raises(ExportError) as raised:
        export_excel(
            workspace,
            document(),
            profile,
            workbook_kind="ut",
            output_directory=tmp_path,
            project="Demo",
            artifact="Login",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_PROFILE_INVALID"


def test_export_revalidates_bypassed_invalid_organization_cell(tmp_path: Path) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT)
    assert profile.organization is not None
    invalid_organization = profile.organization.model_copy(update={"project_cell": "A0"})
    profile = profile.model_copy(update={"organization": invalid_organization})

    with pytest.raises(ExportError) as raised:
        export_excel(
            workspace,
            document(),
            profile,
            workbook_kind="ut",
            output_directory=tmp_path,
            project="Demo",
            artifact="InvalidCell",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_PROFILE_INVALID"
    assert not (tmp_path / "Demo_InvalidCell_Test Case UT.xlsx").exists()


def test_export_revalidates_bypassed_missing_organization_sheet(tmp_path: Path) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT)
    ut = profile.workbooks["ut"]
    without_overview = ut.model_copy(
        update={
            "required_sheets": tuple(
                sheet for sheet in ut.required_sheets if sheet != "Overview"
            )
        }
    )
    profile = profile.model_copy(
        update={"workbooks": {**profile.workbooks, "ut": without_overview}}
    )

    with pytest.raises(ExportError) as raised:
        export_excel(
            workspace,
            document(),
            profile,
            workbook_kind="ut",
            output_directory=tmp_path,
            project="Demo",
            artifact="MissingOverview",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_PROFILE_INVALID"
    assert not (tmp_path / "Demo_MissingOverview_Test Case UT.xlsx").exists()


def test_excel_revalidates_raw_workbooks_container_before_lookup(tmp_path: Path) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT).model_copy(
        update={"workbooks": []}
    )

    with pytest.raises(ExportError) as raised:
        export_excel(
            workspace,
            document(),
            profile,
            workbook_kind="ut",
            output_directory=tmp_path,
            project="Demo",
            artifact="RawContainer",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_PROFILE_INVALID"


def test_excel_revalidates_raw_nested_workbook_before_field_access(tmp_path: Path) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT).model_copy(
        update={"workbooks": {"ut": {"filename": "raw"}}}
    )

    with pytest.raises(ExportError) as raised:
        export_excel(
            workspace,
            document(),
            profile,
            workbook_kind="ut",
            output_directory=tmp_path,
            project="Demo",
            artifact="RawNested",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_PROFILE_INVALID"


def test_markdown_revalidates_profile_before_resource_enumeration(tmp_path: Path) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT).model_copy(
        update={"workbooks": []}
    )

    with pytest.raises(ExportError) as raised:
        export_markdown(
            workspace,
            document(),
            profile,
            tmp_path / "out.md",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_PROFILE_INVALID"
    assert not (tmp_path / "out.md").exists()


def test_output_directory_creation_failure_is_typed(tmp_path: Path) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    blocked_parent = tmp_path / "blocked"
    blocked_parent.write_text("not a directory", encoding="utf-8")

    with pytest.raises(ExportError) as raised:
        export_markdown(
            workspace,
            document(),
            Profile.load("generic", PROFILES_ROOT),
            blocked_parent / "testcases.md",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_WRITE_FAILED"


def test_excel_temporary_file_setup_failure_is_typed(tmp_path: Path, monkeypatch) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT)

    def fail_mkstemp(*args, **kwargs):
        raise OSError("temporary setup failed")

    monkeypatch.setattr(exporter_module.tempfile, "mkstemp", fail_mkstemp)

    with pytest.raises(ExportError) as raised:
        export_excel(
            workspace,
            document(),
            profile,
            workbook_kind="ut",
            output_directory=tmp_path,
            project="Demo",
            artifact="Temp",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_WRITE_FAILED"


def test_excel_temporary_workbook_reload_failure_is_typed(
    tmp_path: Path, monkeypatch
) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT)

    def save_invalid_workbook(workbook, filename) -> None:
        Path(filename).write_bytes(b"not an xlsx zip")

    monkeypatch.setattr(OpenpyxlWorkbook, "save", save_invalid_workbook)

    with pytest.raises(ExportError) as raised:
        export_excel(
            workspace,
            document(),
            profile,
            workbook_kind="ut",
            output_directory=tmp_path,
            project="Demo",
            artifact="Reload",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_WRITE_FAILED"
    assert not (tmp_path / "Demo_Reload_Test Case UT.xlsx").exists()


def test_excel_temporary_workbook_missing_sheet_is_typed(
    tmp_path: Path, monkeypatch
) -> None:
    project = tmp_path / "project"
    project.mkdir()
    workspace = ready_workspace(project)
    profile = Profile.load("company-legacy", PROFILES_ROOT)
    original_save = OpenpyxlWorkbook.save

    def save_without_testcase_sheet(workbook, filename) -> None:
        workbook.remove(workbook["Testcase"])
        original_save(workbook, filename)

    monkeypatch.setattr(OpenpyxlWorkbook, "save", save_without_testcase_sheet)

    with pytest.raises(ExportError) as raised:
        export_excel(
            workspace,
            document(),
            profile,
            workbook_kind="ut",
            output_directory=tmp_path,
            project="Demo",
            artifact="MissingSheet",
            protected_inputs=(),
        )

    assert raised.value.findings[0].code == "EXPORT_WRITE_FAILED"
    assert not (tmp_path / "Demo_MissingSheet_Test Case UT.xlsx").exists()


def test_cli_export_uses_only_explicit_project_cases_profile_and_output_paths(
    tmp_path: Path,
) -> None:
    project = tmp_path / "project"
    project.mkdir()
    ready_workspace(project)
    cases_path = tmp_path / "selected-cases.yaml"
    cases_path.write_text(document().model_dump_json(), encoding="utf-8")
    output = tmp_path / "selected-output.md"

    result = CliRunner().invoke(
        app,
        [
            "export",
            str(project),
            str(cases_path),
            "--profiles-root",
            str(PROFILES_ROOT.resolve()),
            "--profile",
            "generic",
            "--format",
            "markdown",
            "--out",
            str(output),
        ],
    )

    assert result.exit_code == 0
    assert result.stdout.strip() == f"Exported 2 cases to {output}"
    assert output.read_bytes() == render_testcases_markdown(document()).encode("utf-8")


def test_cli_markdown_export_from_canonical_markdown_is_byte_equal(tmp_path: Path) -> None:
    project = tmp_path / "project"
    project.mkdir()
    ready_workspace(project)
    cases = document()
    cases_path = tmp_path / "canonical.md"
    canonical = render_testcases_markdown(cases).encode("utf-8")
    cases_path.write_bytes(canonical)
    output = tmp_path / "exported.md"

    result = CliRunner().invoke(
        app,
        [
            "export",
            str(project),
            str(cases_path),
            "--profiles-root",
            str(PROFILES_ROOT.resolve()),
            "--profile",
            "generic",
            "--format",
            "markdown",
            "--out",
            str(output),
        ],
    )

    assert result.exit_code == 0
    assert output.read_bytes() == canonical


def test_cli_excel_export_parses_canonical_markdown_without_field_loss(tmp_path: Path) -> None:
    project = tmp_path / "project"
    project.mkdir()
    ready_workspace(project)
    cases = document()
    cases.cases[1].tags = ["smoke,critical", "login"]
    cases_path = tmp_path / "canonical.md"
    cases_path.write_text(render_testcases_markdown(cases), encoding="utf-8")
    output_directory = tmp_path / "excel"

    result = CliRunner().invoke(
        app,
        [
            "export",
            str(project),
            str(cases_path),
            "--profiles-root",
            str(PROFILES_ROOT.resolve()),
            "--profile",
            "company-legacy",
            "--format",
            "excel",
            "--out",
            str(output_directory),
            "--workbook",
            "ut",
            "--project-name",
            "Demo",
            "--artifact-name",
            "Canonical",
        ],
    )

    assert result.exit_code == 0
    output = output_directory / "Demo_Canonical_Test Case UT.xlsx"
    sheet = openpyxl.load_workbook(output, data_only=False)["Testcase"]
    assert "Tags: login, smoke,critical" in sheet["D16"].value


def test_cli_unknown_profile_and_format_fail_concisely(tmp_path: Path) -> None:
    project = tmp_path / "project"
    project.mkdir()
    ready_workspace(project)
    cases_path = tmp_path / "cases.yaml"
    cases_path.write_text(document().model_dump_json(), encoding="utf-8")
    runner = CliRunner()

    unknown_profile = runner.invoke(
        app,
        [
            "export",
            str(project),
            str(cases_path),
            "--profiles-root",
            str(PROFILES_ROOT.resolve()),
            "--profile",
            "missing",
            "--format",
            "markdown",
            "--out",
            str(tmp_path / "out.md"),
        ],
    )
    unknown_format = runner.invoke(
        app,
        [
            "export",
            str(project),
            str(cases_path),
            "--profiles-root",
            str(PROFILES_ROOT.resolve()),
            "--profile",
            "generic",
            "--format",
            "pdf",
            "--out",
            str(tmp_path / "out.pdf"),
        ],
    )

    assert unknown_profile.exit_code == 1
    assert "PROFILE_UNKNOWN" in unknown_profile.stdout
    assert unknown_format.exit_code != 0
    assert "Invalid value" in unknown_format.stderr
    assert "Traceback" not in unknown_profile.stdout + unknown_format.stderr
